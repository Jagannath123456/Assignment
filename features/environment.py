from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

def before_all(context):
    # Set up Chrome
    options = webdriver.ChromeOptions()
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--ignore-ssl-errors")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--start-maximized")

    context.driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

    # Set up Excel
    context.excel_file = 'test_results.xlsx'
    context.workbook = openpyxl.Workbook()
    context.sheet = context.workbook.active
    context.sheet.title = "Magento Login Results"
    context.sheet.append(["Email", "Password", "Expected Outcome", "Actual Outcome", "Status"])

def after_scenario(context, scenario):
    # Capture inputs and outcomes
    email = getattr(context, 'email', '')
    password = getattr(context, 'password', '')
    expected = getattr(context, 'expected_outcome', 'Not Set')  # Fix here
    actual = getattr(context, 'actual_outcome', 'Not Set')      # Fix here
    status = getattr(context, 'status', 'FAIL')                 # Optional: capture if you stored it

    # Append to Excel
    row = [email, password, expected, actual, status]
    context.sheet.append(row)

    # Highlight failed test case row in red
    if status == "FAIL":
        fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        for col in range(1, 6):
            context.sheet.cell(row=context.sheet.max_row, column=col).fill = fill


def after_all(context):
    # Save the Excel report with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"magento_login_results_{timestamp}.xlsx"
    context.workbook.save(filename)

    # Close browser
    context.driver.quit
